"""Native .xlsx dashboard export (openpyxl) — the "Tải Excel Dashboard" button
next to the live dashboard. Mirrors the on-screen layout (KPI row + 2-column
chart grid, palette-matched colours) and maps every MiniChart type to the
closest native Excel chart object, so the file still looks like a dashboard
when opened in real Excel — not a data dump with one generic bar chart per
row. Sync/CPU-bound (openpyxl) — callers run this in FastAPI's threadpool.
"""

import io

import openpyxl
from openpyxl.chart import (
    AreaChart, BarChart, BubbleChart, DoughnutChart, LineChart, PieChart,
    RadarChart, Reference, ScatterChart, Series,
)
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Palette accents mirror frontend/src/App.css `--dash-accent` per data-palette,
# so the exported file's colours match whatever the user is looking at on screen.
PALETTES = {
    "emerald": {"accent": "10B981", "series": ["10B981", "0EA5A4", "F59E0B", "EC4899", "8B5CF6", "3B82F6"]},
    "ocean": {"accent": "2563EB", "series": ["2563EB", "0EA5E9", "22C55E", "F59E0B", "8B5CF6", "EC4899"]},
    "sunset": {"accent": "F97316", "series": ["F97316", "EF4444", "F59E0B", "8B5CF6", "22C55E", "2563EB"]},
}

# Chart data is written off-screen (columns Z+) row-banded per chart, same
# trick the Google Sheets builder uses — keeps the visible dashboard clean
# while native chart objects reference real cells (openpyxl charts can't
# embed literal arrays, only cell ranges).
DATA_COL = 26
CHARTS_PER_ROW = 2
CHART_ANCHOR_COLS = ["A", "J"]
CHART_ROW_STEP = 22

# Family assignment: which native Excel chart shape best expresses each of
# the 25 MiniChart.tsx types. Several types have no true Excel equivalent
# (funnel, waterfall, gauge...) and are mapped to the closest readable stand-in
# rather than skipped, per the "never make a chart silently disappear" rule.
BAR_V = {"bar", "lollipop", "dot-plot", "radial-bar", "waterfall"}
BAR_H = {"horizontal-bar", "funnel", "pyramid"}
BAR_H_TARGET = {"gauge", "bullet", "progress"}  # value-vs-target, synthesized as a 2-bar comparison
LINE = {"line", "sparkline", "multi-line"}
AREA = {"area", "stacked-area"}


def _to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _kpi_style(ws, accent: str):
    return {
        "title_font": Font(name="Segoe UI", size=16, bold=True, color=accent),
        "kpi_title_font": Font(name="Segoe UI", size=9, bold=True, color="4B5563"),
        "kpi_val_font": Font(name="Segoe UI", size=18, bold=True, color=accent),
        "kpi_fill": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
        "chart_title_font": Font(name="Segoe UI", size=12, bold=True, color="1F2937"),
        "border": Border(
            left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"),
        ),
    }


def _write_kpi(ws, style, col: int, title: str, value) -> None:
    c1, c2 = get_column_letter(col), get_column_letter(col + 1)
    ws.merge_cells(f"{c1}3:{c2}3")
    ws.merge_cells(f"{c1}4:{c2}4")

    t = ws[f"{c1}3"]
    t.value = title
    t.font = style["kpi_title_font"]
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill = style["kpi_fill"]

    v = ws[f"{c1}4"]
    try:
        v.value = float(value) if isinstance(value, (int, float, str)) and str(value).replace(".", "", 1).lstrip("-").isdigit() else value
    except Exception:
        v.value = value
    v.font = style["kpi_val_font"]
    v.alignment = Alignment(horizontal="center", vertical="center")
    v.fill = style["kpi_fill"]

    for r in (3, 4):
        for c in (col, col + 1):
            ws.cell(row=r, column=c).border = style["border"]


def _write_category_block(ws, row0: int, labels: list, series: list[dict]) -> tuple[int, int]:
    """categories in col DATA_COL, one value column per series starting at
    DATA_COL+1, header row = series names. Returns (n_rows, n_series)."""
    ws.cell(row=row0, column=DATA_COL, value="Nhãn")
    for si, s in enumerate(series):
        ws.cell(row=row0, column=DATA_COL + 1 + si, value=s.get("name") or f"Series {si + 1}")
    for i, label in enumerate(labels):
        ws.cell(row=row0 + 1 + i, column=DATA_COL, value=label)
        for si, s in enumerate(series):
            vals = s.get("values") or []
            ws.cell(row=row0 + 1 + i, column=DATA_COL + 1 + si, value=_to_num(vals[i]) if i < len(vals) else 0)
    return len(labels), len(series)


def _write_point_block(ws, row0: int, points: list[dict], bubble: bool) -> int:
    ws.cell(row=row0, column=DATA_COL, value="X")
    ws.cell(row=row0, column=DATA_COL + 1, value="Y")
    if bubble:
        ws.cell(row=row0, column=DATA_COL + 2, value="Size")
    for i, p in enumerate(points):
        ws.cell(row=row0 + 1 + i, column=DATA_COL, value=_to_num(p.get("x")))
        ws.cell(row=row0 + 1 + i, column=DATA_COL + 1, value=_to_num(p.get("y")))
        if bubble:
            ws.cell(row=row0 + 1 + i, column=DATA_COL + 2, value=_to_num(p.get("size", 1)) or 1)
    return len(points)


def _add_category_chart(ws, chart_type: str, title: str, row0: int, n_rows: int, n_series: int,
                         colors: list[str], style, grouping: str | None = None) -> object:
    if chart_type == "pie":
        chart = PieChart()
    elif chart_type == "donut":
        chart = DoughnutChart()
    elif chart_type in LINE:
        chart = LineChart()
    elif chart_type in AREA:
        chart = AreaChart()
        if chart_type == "stacked-area":
            chart.grouping = "stacked"
    elif chart_type == "radar":
        chart = RadarChart()
    else:
        chart = BarChart()
        chart.type = "bar" if chart_type in BAR_H or chart_type in BAR_H_TARGET else "col"
        if grouping:
            chart.grouping = grouping
            chart.overlap = -10 if grouping == "clustered" else 100

    chart.title = title
    chart.style = 10
    chart.width, chart.height = 16, 10

    data_ref = Reference(ws, min_col=DATA_COL + 1, max_col=DATA_COL + n_series, min_row=row0, max_row=row0 + n_rows)
    cats_ref = Reference(ws, min_col=DATA_COL, min_row=row0 + 1, max_row=row0 + n_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    for i, s in enumerate(chart.series):
        hexc = colors[i % len(colors)]
        if chart_type in ("pie", "donut"):
            continue  # per-slice colour, not per-series — leave Excel's default palette
        if chart_type in LINE:
            s.smooth = False
            s.graphicalProperties.line.solidFill = hexc
            s.graphicalProperties.line.width = 20000
            s.marker = Marker(symbol="circle", size=5)
            s.marker.graphicalProperties.solidFill = hexc
        else:
            s.graphicalProperties.solidFill = hexc

    if chart_type not in ("pie", "donut") and n_series <= 1:
        chart.legend = None

    return chart


def _add_scatter_chart(ws, title: str, row0: int, n_rows: int, bubble: bool, color: str) -> object:
    chart = BubbleChart() if bubble else ScatterChart()
    chart.title = title
    chart.style = 13
    chart.width, chart.height = 16, 10
    x_ref = Reference(ws, min_col=DATA_COL, min_row=row0 + 1, max_row=row0 + n_rows)
    y_ref = Reference(ws, min_col=DATA_COL + 1, min_row=row0 + 1, max_row=row0 + n_rows)
    if bubble:
        size_ref = Reference(ws, min_col=DATA_COL + 2, min_row=row0 + 1, max_row=row0 + n_rows)
        series = Series(y_ref, x_ref, zvalues=size_ref, title=title)
    else:
        series = Series(y_ref, x_ref, title=title)
        series.marker = Marker(symbol="circle", size=6)
        series.graphicalProperties.line.noFill = True
    series.graphicalProperties.solidFill = color
    chart.series.append(series)
    chart.legend = None
    return chart


def _add_heatmap(ws, style, title: str, row0: int, labels: list, row_labels: list, matrix: list[list]) -> int:
    """No native "heatmap chart" exists in Excel — a conditional-formatting
    colour scale over a real grid IS the idiomatic Excel heatmap, and reads
    better than faking it with a chart object."""
    ws.cell(row=row0, column=DATA_COL, value=title).font = style["chart_title_font"]
    header_row = row0 + 1
    for ci, lbl in enumerate(labels):
        ws.cell(row=header_row, column=DATA_COL + 1 + ci, value=lbl).font = Font(bold=True, size=9)
    for ri, row in enumerate(matrix):
        ws.cell(row=header_row + 1 + ri, column=DATA_COL, value=row_labels[ri] if ri < len(row_labels) else "").font = Font(bold=True, size=9)
        for ci, v in enumerate(row):
            ws.cell(row=header_row + 1 + ri, column=DATA_COL + 1 + ci, value=_to_num(v))

    n_rows, n_cols = len(matrix), len(labels)
    if n_rows and n_cols:
        rng = (
            f"{get_column_letter(DATA_COL + 1)}{header_row + 1}:"
            f"{get_column_letter(DATA_COL + n_cols)}{header_row + n_rows}"
        )
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="FFF7ED",
                mid_type="percentile", mid_value=50, mid_color="FDBA74",
                end_type="max", end_color="EA580C",
            ),
        )
    return n_rows + 3


def _add_chart_item(ws, style, item: dict, anchor: str, row0: int, colors: list[str], color_index: int) -> int:
    """Writes one dashboard chart item (data block + native chart/heatmap grid
    anchored at `anchor`). Returns the number of data rows consumed, so the
    caller can size the next chart's off-screen row band."""
    chart_spec = item.get("chart") or {}
    title = chart_spec.get("title") or item.get("title", "")
    ctype = chart_spec.get("type", "bar")
    labels = chart_spec.get("labels") or []
    values = chart_spec.get("values") or []
    series = chart_spec.get("series")
    points = chart_spec.get("points")
    matrix = chart_spec.get("matrix")
    accent = colors[color_index % len(colors)]

    if ctype == "heatmap" and matrix:
        return _add_heatmap(ws, style, title, row0, labels, chart_spec.get("rowLabels") or [], matrix)

    if ctype in ("scatter", "bubble") and points:
        n = _write_point_block(ws, row0, points, bubble=(ctype == "bubble"))
        if n:
            chart = _add_scatter_chart(ws, title, row0, n, bubble=(ctype == "bubble"), color=accent)
            ws.add_chart(chart, anchor)
        return n + 2

    if ctype in BAR_H_TARGET:
        val = values[0] if values else 0
        target = chart_spec.get("target") if chart_spec.get("target") is not None else chart_spec.get("max")
        synth_labels = ["Đạt được", "Mục tiêu" if chart_spec.get("target") is not None else "Tối đa"]
        n_rows, n_series = _write_category_block(ws, row0, synth_labels, [{"name": title, "values": [val, target or 0]}])
        chart = _add_category_chart(ws, "horizontal-bar", title, row0, n_rows, n_series, [accent], style)
        ws.add_chart(chart, anchor)
        return n_rows + 2

    if series:
        grouping = "stacked" if ctype == "stacked-bar" else ("clustered" if ctype == "grouped-bar" else None)
        n_rows, n_series = _write_category_block(ws, row0, labels, series)
        if ctype == "combo" and len(series) > 1:
            bar_rows, bar_n = _write_category_block(ws, row0, labels, series[:1])
            bar_chart = _add_category_chart(ws, "bar", title, row0, bar_rows, bar_n, colors, style)
            line_chart = LineChart()
            data_ref = Reference(ws, min_col=DATA_COL + 2, max_col=DATA_COL + 1 + len(series), min_row=row0, max_row=row0 + n_rows)
            cats_ref = Reference(ws, min_col=DATA_COL, min_row=row0 + 1, max_row=row0 + n_rows)
            line_chart.add_data(data_ref, titles_from_data=True)
            line_chart.set_categories(cats_ref)
            for i, s in enumerate(line_chart.series):
                s.graphicalProperties.line.solidFill = colors[(i + 1) % len(colors)]
                s.graphicalProperties.line.width = 20000
            bar_chart += line_chart
            bar_chart.title = title
            ws.add_chart(bar_chart, anchor)
        else:
            chart = _add_category_chart(ws, ctype, title, row0, n_rows, n_series, colors, style, grouping=grouping)
            ws.add_chart(chart, anchor)
        return n_rows + 2

    if not values:
        return 2
    n_rows, n_series = _write_category_block(ws, row0, labels, [{"name": title, "values": values}])
    chart = _add_category_chart(ws, ctype, title, row0, n_rows, n_series, colors, style)
    ws.add_chart(chart, anchor)
    return n_rows + 2


def build_dashboard_xlsx(items: list[dict], palette: str = "emerald") -> io.BytesIO:
    colors = PALETTES.get(palette, PALETTES["emerald"])
    accent, series_colors = colors["accent"], colors["series"]

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    ws = wb.create_sheet(title="Dashboard", index=0)
    ws.sheet_view.showGridLines = False

    style = _kpi_style(ws, accent)
    ws["A1"] = "BÁO CÁO PHÂN TÍCH DOANH NGHIỆP"
    ws["A1"].font = style["title_font"]
    ws.row_dimensions[1].height = 24

    for col in range(1, 40):
        ws.column_dimensions[get_column_letter(col)].width = 15

    kpi_col = 1
    chart_row = 6
    slot = 0
    color_idx = 0

    for item in items:
        itype = item.get("type")
        if itype == "kpi":
            _write_kpi(ws, style, kpi_col, item.get("title", ""), item.get("scalar"))
            kpi_col += 3
        elif itype == "chart":
            anchor_col = CHART_ANCHOR_COLS[slot % CHARTS_PER_ROW]
            anchor_row = chart_row + (slot // CHARTS_PER_ROW) * CHART_ROW_STEP
            data_row0 = 200 + slot * 60  # off-screen band, generous headroom per chart
            consumed = _add_chart_item(ws, style, item, f"{anchor_col}{anchor_row}", data_row0, series_colors, color_idx)
            slot += 1
            color_idx += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
